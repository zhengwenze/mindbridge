from __future__ import annotations

import smtplib
import ssl
import threading
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.core.enums import RiskCaseStatus, ToolStatus
from app.models.entities import AlertRecord, CaseNote, ExcelRecord, PsychologicalReport, RiskCase, UserAccount
from app.services.skills import MindBridgeSkillLibrary


EXCEL_WRITE_LOCK = threading.Lock()


class ToolOrchestrationService:
    def __init__(self, db: Session, settings: Settings):
        self.db = db
        self.settings = settings

    def write_excel(self, report: PsychologicalReport) -> ExcelRecord:
        existing = (
            self.db.query(ExcelRecord)
            .filter(ExcelRecord.report_id == report.id, ExcelRecord.status == ToolStatus.SUCCESS.value)
            .first()
        )
        if existing is not None:
            return existing
        path = Path(self.settings.excel_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with EXCEL_WRITE_LOCK:
            if path.exists():
                workbook = load_workbook(path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "MindBridge Risk Ledger"
                sheet.append(["reportId", "riskLevel", "emotion", "confidence", "summary", "createdAt"])
            sheet.append([report.id, report.risk_level, report.emotion, report.confidence, report.summary, report.created_at.isoformat()])
            workbook.save(path)
        record = ExcelRecord(report_id=report.id, file_path=str(path), status=ToolStatus.SUCCESS.value, message="Excel 台账已写入")
        self.db.add(record)
        self.db.commit()
        return record

    def create_case(self, report: PsychologicalReport) -> RiskCase:
        existing = self.db.query(RiskCase).filter(RiskCase.report_id == report.id).first()
        if existing is not None:
            return existing
        user = self.db.get(UserAccount, report.user_id)
        case = RiskCase(
            report_id=report.id,
            risk_level=report.risk_level,
            status=RiskCaseStatus.OPEN.value,
            owner=self._primary_owner(),
            summary=report.summary,
            handoff_summary=MindBridgeSkillLibrary.counselor_handoff_summary(report, user),
        )
        self.db.add(case)
        self.db.commit()
        return case

    def send_case_alert(self, case: RiskCase) -> AlertRecord:
        report = self.db.get(PsychologicalReport, case.report_id)
        if report is None:
            raise RuntimeError(f"report {case.report_id} not found")
        record = self.notify(report, case)
        if record.status == ToolStatus.SUCCESS.value and case.status == RiskCaseStatus.OPEN.value:
            case.status = RiskCaseStatus.ALERT_SENT.value
        case.updated_at = datetime.utcnow()
        self.db.add(case)
        self.db.commit()
        return record

    def acknowledge_case(self, case_id: int, actor: str, note: str = "") -> RiskCase:
        case = self.db.get(RiskCase, case_id)
        if case is None:
            raise RuntimeError(f"case {case_id} not found")
        actor_name = actor.strip() or "unknown"
        case.status = RiskCaseStatus.ACKNOWLEDGED.value
        case.acknowledged_by = actor_name
        case.acknowledged_at = datetime.utcnow()
        case.updated_at = datetime.utcnow()
        self.db.add(case)
        self._add_case_note(case.id, actor_name, note.strip() or "已确认接手该个案")
        self.db.commit()
        return case

    def add_case_note(self, case_id: int, actor: str, note: str) -> CaseNote:
        case = self.db.get(RiskCase, case_id)
        if case is None:
            raise RuntimeError(f"case {case_id} not found")
        case.updated_at = datetime.utcnow()
        self.db.add(case)
        record = self._add_case_note(case.id, actor.strip() or "unknown", note.strip())
        self.db.commit()
        return record

    def notify(self, report: PsychologicalReport, case: RiskCase | None = None) -> AlertRecord:
        existing = (
            self.db.query(AlertRecord)
            .filter(AlertRecord.report_id == report.id, AlertRecord.status == ToolStatus.SUCCESS.value)
            .first()
        )
        if existing is not None:
            return existing
        recipient = self.settings.alert_email_to.strip() or "unconfigured"
        mode = self.settings.alert_email_delivery_mode.strip().lower()
        if mode == "log":
            return self._save_alert(
                report,
                recipient if recipient != "unconfigured" else "log",
                ToolStatus.SUCCESS.value,
                f"高风险预警已记录：reportId={report.id}，caseId={case.id if case else 'none'}，deliveryMode=log",
            )
        if mode != "smtp":
            return self._save_alert(
                report,
                recipient,
                ToolStatus.FAILED.value,
                f"高风险预警邮件未发送：未知投递模式 {self.settings.alert_email_delivery_mode}",
            )
        missing = self._missing_email_config()
        if missing:
            return self._save_alert(
                report,
                recipient,
                ToolStatus.FAILED.value,
                f"高风险预警邮件未发送：缺少配置 {', '.join(missing)}",
            )
        try:
            self._send_alert_email(report, case)
        except Exception as exc:
            return self._save_alert(
                report,
                recipient,
                ToolStatus.FAILED.value,
                f"高风险预警邮件发送失败：{type(exc).__name__}: {exc}",
            )
        return self._save_alert(report, recipient, ToolStatus.SUCCESS.value, f"高风险预警邮件已发送：reportId={report.id}")

    def _save_alert(self, report: PsychologicalReport, recipient: str, status: str, message: str) -> AlertRecord:
        record = AlertRecord(
            report_id=report.id,
            channel="email",
            recipient=recipient,
            status=status,
            message=message,
        )
        self.db.add(record)
        self.db.commit()
        return record

    def _add_case_note(self, case_id: int, actor: str, note: str) -> CaseNote:
        if not note:
            raise RuntimeError("case note cannot be empty")
        record = CaseNote(case_id=case_id, actor=actor, note=note)
        self.db.add(record)
        return record

    def _missing_email_config(self) -> list[str]:
        missing = []
        if not self.settings.smtp_host.strip():
            missing.append("SMTP_HOST")
        if not self._sender():
            missing.append("ALERT_EMAIL_FROM 或 SMTP_USERNAME")
        if not self._recipients():
            missing.append("ALERT_EMAIL_TO")
        return missing

    def _send_alert_email(self, report: PsychologicalReport, case: RiskCase | None = None) -> None:
        message = EmailMessage()
        message["Subject"] = f"{self.settings.alert_email_subject_prefix} reportId={report.id}"
        message["From"] = self._sender()
        message["To"] = ", ".join(self._recipients())
        message.set_content(self._email_body(report, case))

        context = ssl.create_default_context()
        if self.settings.smtp_use_ssl:
            with smtplib.SMTP_SSL(
                self.settings.smtp_host,
                self.settings.smtp_port,
                timeout=self.settings.smtp_timeout_seconds,
                context=context,
            ) as server:
                self._send_message(server, message)
            return

        with smtplib.SMTP(self.settings.smtp_host, self.settings.smtp_port, timeout=self.settings.smtp_timeout_seconds) as server:
            server.ehlo()
            if self.settings.smtp_use_tls:
                server.starttls(context=context)
                server.ehlo()
            self._send_message(server, message)

    def _send_message(self, server: smtplib.SMTP, message: EmailMessage) -> None:
        if self.settings.smtp_username:
            server.login(self.settings.smtp_username, self.settings.smtp_password)
        server.send_message(message)

    def _email_body(self, report: PsychologicalReport, case: RiskCase | None = None) -> str:
        user = self.db.get(UserAccount, report.user_id)
        username = user.username if user else f"userId={report.user_id}"
        display_name = user.display_name if user else ""
        handoff = case.handoff_summary if case else MindBridgeSkillLibrary.counselor_handoff_summary(report, user)
        return "\n".join(
            [
                "MindBridge 检测到一条高风险心理预警，请尽快安排辅导员或管理员跟进。",
                "",
                f"个案ID：{case.id}" if case else "个案ID：未创建",
                f"报告ID：{report.id}",
                f"学生：{display_name} ({username})" if display_name else f"学生：{username}",
                f"风险等级：{report.risk_level}",
                f"情绪标签：{report.emotion}",
                f"置信度：{report.confidence}",
                f"摘要：{report.summary}",
                f"创建时间：{report.created_at.isoformat()}",
                "",
                "交接摘要：",
                handoff,
            ]
        )

    def _primary_owner(self) -> str:
        recipients = self._recipients()
        if recipients:
            return recipients[0]
        return "unassigned"

    def _sender(self) -> str:
        return self.settings.alert_email_from.strip() or self.settings.smtp_username.strip()

    def _recipients(self) -> list[str]:
        normalized = self.settings.alert_email_to.replace(";", ",")
        return [recipient.strip() for recipient in normalized.split(",") if recipient.strip()]
